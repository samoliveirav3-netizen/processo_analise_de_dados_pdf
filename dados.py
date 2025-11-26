import tkinter as tk
import pandas as pd
import customtkinter
from openpyxl import load_workbook
from tkinter import filedialog, messagebox

#Crie o script para abrir a planilha
dados = pd.read_excel('DATASETE/dados.xlsx')
df = pd.DataFrame(dados)
print(dados)

#Crie o script para acessar o arquivo
ar = load_workbook('DATASETE/dados.xlsx')
ar2 = ar.active # Selecionar a aba ativa
celula_1 = ar2.cell(row=4, column=2)
print(celula_1.value)

# def dados_da_planilha():
#     # Abre uma caixa de diálogo para selecionar o arquivo
#     caminho_arquivo = filedialog.askopenfilename(
#         filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Arquivos CSV", "*.csv")]
#     )

#     if caminho_arquivo:
#         try:
#             if caminho_arquivo.endswith('.csv'):
#                 # Lê o arquivo CSV usando pandas
#                 df = pd.read_csv(caminho_arquivo)
#             else:
#                 # Lê o arquivo Excel usando pandas
#                 df = pd.read_excel(caminho_arquivo)
            
#             # Exemplo de processamento ou exibição dos dados (aqui apenas imprime no console)
#             print("Dados carregados com sucesso:")
#             print(df.head()) # Mostra as primeiras 5 linhas

#             # Aqui você pode adicionar a lógica para exibir os dados em um widget Treeview, por exemplo
#             messagebox.showinfo("Sucesso", f"Planilha carregada: {caminho_arquivo}\nTotal de linhas: {len(df)}")

#         except Exception as e:
#             messagebox.showerror("Erro ao carregar", f"Ocorreu um erro ao ler o arquivo: {e}")

#Crie uma Janela para mostrar os dados da Planilha 
estastics = df.describe()
janela = tk.Tk()
a = tk.Label(janela,text=estastics).pack()
janela.geometry('750x550')
janela.configure(bg = 'black')
janela.title()

tk.Label(janela,text='GRAFICO DE BARRA', font=('arial',12), bg= 'black', fg = 'white').pack(pady=10)
customtkinter.CTkButton(janela, text='Grafico Excel(Barra)', font=('arial',12), command=a).pack(pady=10)

janela.mainloop()

